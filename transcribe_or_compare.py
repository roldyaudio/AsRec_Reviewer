import argparse
import contextlib
import os
import sys
import time
from collections import Counter
from datetime import datetime
from pathlib import Path

import jiwer
from dataclasses import dataclass
from typing import Dict, Iterable, List, Optional, TextIO, Union

from deepgram import DeepgramClient, PrerecordedOptions
import tempfile
from concurrent.futures import ThreadPoolExecutor, as_completed
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from rpp_qa import (
    QAProjectConfig,
    QA_COLOR_GREEN,
    QA_COLOR_RED,
    QA_COLOR_YELLOW,
    generate_qa_project,
)


# Colores suaves compartidos con el proyecto QA de REAPER.
def excel_fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


GREEN_FILL = excel_fill(QA_COLOR_GREEN)
RED_FILL = excel_fill(QA_COLOR_RED)
YELLOW_FILL = excel_fill(QA_COLOR_YELLOW)

# STT quality thresholds. Lower error rates are better.
WER_EXCELLENT_THRESHOLD = 0.05
WER_ACCEPTABLE_THRESHOLD = 0.15
CER_EXCELLENT_THRESHOLD = 0.02
CER_ACCEPTABLE_THRESHOLD = 0.08

ASIAN_LANGUAGES = {"JP", "JA", "KO", "KR", "TH", "CN", "ZH", "ZH-CN", "ZH-TW"}
QUALITY_EXCELLENT = "excellent"
QUALITY_ACCEPTABLE = "acceptable"
QUALITY_POOR = "poor"
QUALITY_UNEVALUATED = "not_evaluated"
STATUS_OK = "ok"
STATUS_FAILED = "failed"
STATUS_MISSING_AUDIO = "missing_audio"
STATUS_NO_EXCEL_MATCH = "no_excel_match"
STATUS_EMPTY_TRANSCRIPT = "empty_transcript"
DEFAULT_DEEPGRAM_RETRIES = 3
DEFAULT_DEEPGRAM_RETRY_BACKOFF_SECONDS = 2.0

wer_transforms = jiwer.Compose([
    jiwer.ToLowerCase(),
    jiwer.RemovePunctuation(),
    jiwer.SubstituteRegexes({r"[^\w\s]": ""}),
    jiwer.RemoveMultipleSpaces(),
    jiwer.Strip(),
])

cer_transforms = jiwer.Compose([
    jiwer.Strip(),
])


class BaseTranscriber:
    """Interfaz base para cambiar facilmente el backend de transcripcion."""

    def transcribe(self, audio_path: str, language: Optional[str] = None) -> str:
        raise NotImplementedError


class _TeeStream:
    """Mirror stdout/stderr to the terminal and to a persistent run log."""

    def __init__(self, *streams: TextIO):
        self.streams = streams

    def write(self, data: str) -> int:
        for stream in self.streams:
            stream.write(data)
            stream.flush()
        return len(data)

    def flush(self) -> None:
        for stream in self.streams:
            stream.flush()


def default_log_path(output_path: Optional[str] = None) -> Path:
    """Return a timestamped log path near the output file when possible."""
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    directory = Path(output_path).expanduser().parent if output_path else Path.cwd()
    if str(directory) in {"", "."}:
        directory = Path.cwd()
    directory.mkdir(parents=True, exist_ok=True)
    return directory / f"asrec_run_{timestamp}.log"


@contextlib.contextmanager
def run_logging(log_path: Optional[str] = None, output_path: Optional[str] = None):
    """Write all console output to a log file without hiding it from the user."""
    path = Path(log_path).expanduser() if log_path else default_log_path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    original_stdout = sys.stdout
    original_stderr = sys.stderr
    with path.open("a", encoding="utf-8") as log_file:
        sys.stdout = _TeeStream(original_stdout, log_file)
        sys.stderr = _TeeStream(original_stderr, log_file)
        print(f"[INFO] Log de ejecución: {path}")
        try:
            yield path
        finally:
            sys.stdout = original_stdout
            sys.stderr = original_stderr




class DeepgramTranscriber(BaseTranscriber):
    """Implementación usando Deepgram API (nova-3)."""

    def __init__(
        self,
        api_key: str,
        model: str = "nova-3",
        max_workers: int = 4,
        glossary_path: Optional[str] = None,
        max_retries: int = DEFAULT_DEEPGRAM_RETRIES,
        retry_backoff_seconds: float = DEFAULT_DEEPGRAM_RETRY_BACKOFF_SECONDS,
    ):
        if not api_key:
            raise ValueError("Deepgram API key no configurada. Define DEEPGRAM_API_KEY.")

        self.api_key = api_key
        self.model = model
        self.max_workers = max(1, int(max_workers))
        self.max_retries = max(1, int(max_retries))
        self.retry_backoff_seconds = max(0.1, float(retry_backoff_seconds))
        self.keyterms = self._load_keyterms_from_glossary(glossary_path)

        print("[INFO] Inicializando Deepgram...")
        print(f"[INFO] Modelo Deepgram: {model}")
        print(f"[INFO] Workers Deepgram: {self.max_workers}")
        print(f"[INFO] Reintentos Deepgram por audio: {self.max_retries}")
        self._print_keyterms_summary()

        self.client = DeepgramClient(api_key)

    @staticmethod
    def _is_retryable_error(exc: Exception) -> bool:
        text = str(exc).lower()
        retryable_markers = [
            "timeout",
            "timed out",
            "temporarily unavailable",
            "connection",
            "reset by peer",
            "429",
            "rate limit",
            "too many requests",
            "500",
            "502",
            "503",
            "504",
            "gateway",
            "service unavailable",
        ]
        return any(marker in text for marker in retryable_markers)

    @staticmethod
    def _is_enabled(value) -> bool:
        if pd.isna(value):
            return False
        if isinstance(value, bool):
            return value
        text = str(value).strip().lower()
        return text in {"1", "true", "t", "yes", "y", "si", "sí"}

    @staticmethod
    def _split_variants(value) -> List[str]:
        if pd.isna(value):
            return []
        return [part.strip() for part in str(value).split(",") if part.strip()]

    def _load_keyterms_from_glossary(self, glossary_path: Optional[str]) -> List[str]:
        if not glossary_path:
            return []
        if not os.path.exists(glossary_path):
            raise FileNotFoundError(f"No existe el glosario: {glossary_path}")

        glossary_df = pd.read_excel(glossary_path)
        expected_columns = {"term", "boost", "variants", "enabled", "notes"}
        missing = expected_columns - set(glossary_df.columns)
        if missing:
            raise ValueError(
                f"El glosario debe incluir columnas {sorted(expected_columns)}. "
                f"Faltantes: {sorted(missing)}"
            )

        keyterms: List[str] = []
        for _, row in glossary_df.iterrows():
            if not self._is_enabled(row.get("enabled")):
                continue

            term = str(row.get("term", "")).strip()
            if not term:
                continue

            tokens = [term, *self._split_variants(row.get("variants"))]
            for token in tokens:
                keyterms.append(token)

        # Mantener orden original removiendo duplicados exactos.
        return list(dict.fromkeys(keyterms))

    def _print_keyterms_summary(self) -> None:
        total_keyterms = len(self.keyterms)
        if total_keyterms == 0:
            print("[INFO] Glosario: sin keyterms habilitados (se enviará Deepgram sin keyterms).")
            return

        if 10 <= total_keyterms <= 50:
            level = "🟢 ideal"
        elif 51 <= total_keyterms <= 120:
            level = "🟡 ok"
        else:
            level = "🔴 riesgo"

        print(f"[INFO] Keyterms Deepgram cargados: {total_keyterms}")
        print("[INFO] Nivel recomendado de keyterms reales:")
        print("       🟢 ideal: 10 – 50")
        print("       🟡 ok: 50 – 120")
        print("       🔴 riesgo: 120+")
        print(f"[INFO] Nivel actual: {level}")

    def transcribe(self, audio_path: str, language: Optional[str] = None) -> str:
        payload: Dict[str, bytes]
        with open(audio_path, "rb") as file:
            payload = {"buffer": file.read()}

        options = PrerecordedOptions(
            model=self.model,
            smart_format=True,
            language=language if language else None,
            keyterm=self.keyterms if self.keyterms else None,
        )

        last_error: Optional[Exception] = None
        for attempt in range(1, self.max_retries + 1):
            try:
                response = self.client.listen.rest.v("1").transcribe_file(payload, options)
                transcript = response.results.channels[0].alternatives[0].transcript
                return (transcript or "").strip()
            except Exception as exc:
                last_error = exc
                retryable = self._is_retryable_error(exc)
                if attempt >= self.max_retries or not retryable:
                    raise
                wait_seconds = self.retry_backoff_seconds * (2 ** (attempt - 1))
                print(
                    f"[WARN] Deepgram intento {attempt}/{self.max_retries} falló para "
                    f"{os.path.basename(audio_path)}: {exc}. Reintentando en {wait_seconds:.1f}s..."
                )
                time.sleep(wait_seconds)

        if last_error:
            raise last_error
        return ""

class WhisperTranscriber(BaseTranscriber):
    """Implementacion con OpenAI Whisper local."""

    def __init__(self, model_size: str = "medium"):
        import torch
        import whisper

        self.device = "cuda" if torch.cuda.is_available() else "cpu"
        self.model_size = model_size
        
        print(f"[INFO] Usando dispositivo: {self.device}")
        print(f"[INFO] Cargando modelo Whisper: {model_size}")
        
        # Whisper cargará automáticamente large-v3 si pasas ese string
        self.model = whisper.load_model(model_size, device=self.device)

    def transcribe(self, audio_path: str, language: Optional[str] = None) -> str:
        from pydub import AudioSegment
        from pydub.silence import split_on_silence

        kwargs = {"fp16": self.device == "cuda"}
        if language:
            kwargs["language"] = language

        # 1. Obtener duración para decidir si segmentar
        audio = AudioSegment.from_file(audio_path)
        duration_secs = len(audio) / 1000.0

        # 2. Si el audio es corto (ej. menos de 15s), transcribir directo
        if duration_secs < 15.0:
            result = self.model.transcribe(audio_path, **kwargs)
            return result.get("text", "").strip()

        # 3. Solo si es un audio largo, usar la lógica de chunks (opcional)
        chunks = split_on_silence(
            audio,
            min_silence_len=700, # Aumentado para evitar micro-segmentos
            silence_thresh=-45,
            keep_silence=500     # Más margen ayuda a Whisper a entender el inicio/fin
        )
        
        print(f"[DEBUG] Numero de chunks: {len(chunks)}")
        
        if not chunks:
            print("[WARN] Silence split fallo, usando transcripcion directa.")
            result = self.model.transcribe(audio_path, **kwargs)
            return result.get("text", "").strip()

        full_text = []
        print(f"[INFO] Procesando {len(chunks)} chunks...")

        for i, chunk in enumerate(chunks):
            # 1. Validacion previa: Si el chunk es muy corto, ni siquiera creamos el temporal
            if len(chunk) < 500:
                continue

            # 2. Crear el archivo temporal pero con delete=False para manejarlo nosotros
            tmp = tempfile.NamedTemporaryFile(suffix=".wav", delete=False)
            try:
                # Exportar el audio al archivo temporal
                chunk.export(tmp.name, format="wav")
                
                # 3. IMPORTANTE: Cerramos el puntero del archivo para que Windows/Linux 
                # permitan que Whisper lo abra sin bloqueos.
                tmp.close()

                # Transcribir usando la ruta del archivo
                result = self.model.transcribe(tmp.name, **kwargs)
                text = result.get("text", "").strip()

                if text:
                    full_text.append(text)

            except Exception as e:
                print(f"[ERROR] Chunk {i} fallo: {e}")

            finally:
                # 4. Limpieza garantizada: Si el archivo existe, se borra si o si.
                if os.path.exists(tmp.name):
                    try:
                        os.remove(tmp.name)
                    except Exception as e:
                        print(f"[WARN] No se pudo borrar temporal {tmp.name}: {e}")
        print(f"Proceso completado")
        return " ".join(full_text).strip()


@dataclass
class ComparisonResult:
    audio_file: str
    transcript: str
    expected_text: str
    wer: Optional[float]
    cer: Optional[float]
    quality: str
    transcription_status: str = STATUS_OK
    transcription_error: str = ""


@dataclass
class TranscriptionOnlyResult:
    audio_file: str
    transcript: str
    transcription_status: str = STATUS_OK
    transcription_error: str = ""


@dataclass
class TranscriptionRecord:
    audio_file: str
    transcript: str = ""
    status: str = STATUS_OK
    error: str = ""


@dataclass
class ValidationSummary:
    audio_count: int
    excel_rows: int = 0
    exact_matches: int = 0
    missing_in_audio_folder: int = 0
    audio_without_excel_match: int = 0
    match_ratio: float = 0.0


TranscriptMap = Dict[str, Union[str, TranscriptionRecord]]


def get_audio_files(folder_path: str, extensions: Iterable[str] = (".wav", ".mp3", ".m4a")) -> List[str]:
    audio_files: List[str] = []
    allowed = tuple(ext.lower() for ext in extensions)

    for root, _, files in os.walk(folder_path):
        for file in files:
            if file.lower().endswith(allowed):
                audio_files.append(os.path.join(root, file))

    audio_files.sort()
    return audio_files


def _record_from_value(audio_file: str, value: Union[str, TranscriptionRecord, None]) -> TranscriptionRecord:
    if isinstance(value, TranscriptionRecord):
        return value
    return TranscriptionRecord(audio_file=audio_file, transcript=value or "")


def transcript_text(value: Union[str, TranscriptionRecord, None]) -> str:
    return _record_from_value("", value).transcript


def _excel_audio_values(
    excel_path: str,
    audio_column: str,
    sheet_name: Optional[str] = None,
) -> List[str]:
    df = pd.read_excel(excel_path, sheet_name=sheet_name)
    if isinstance(df, dict):
        if len(df) > 1:
            print("[WARN] Se detectaron multiples hojas, usando la primera.")
        df = list(df.values())[0]
    if audio_column not in df.columns:
        raise ValueError(
            f"Columna de audio invalida. Encontradas: {list(df.columns)} | "
            f"Esperada: '{audio_column}'"
        )
    return [str(value).strip() for value in df[audio_column].fillna("").tolist() if str(value).strip()]


def validate_inputs(
    audio_folder: str,
    excel_path: Optional[str] = None,
    audio_column: str = "Filename",
    sheet_name: Optional[str] = None,
) -> ValidationSummary:
    """Print and return a preflight summary before spending time/API quota."""
    audio_files = get_audio_files(audio_folder)
    relative_audio = {os.path.relpath(path, audio_folder) for path in audio_files}
    summary = ValidationSummary(audio_count=len(audio_files))

    if excel_path:
        excel_audio = _excel_audio_values(excel_path, audio_column, sheet_name)
        excel_audio_set = set(excel_audio)
        summary.excel_rows = len(excel_audio)
        summary.exact_matches = sum(1 for audio_name in excel_audio if audio_name in relative_audio)
        summary.missing_in_audio_folder = sum(1 for audio_name in excel_audio if audio_name not in relative_audio)
        summary.audio_without_excel_match = len(relative_audio - excel_audio_set)
        summary.match_ratio = summary.exact_matches / summary.excel_rows if summary.excel_rows else 0.0

    print("\n=== Validación previa ===")
    print(f"Audios detectados: {summary.audio_count}")
    if excel_path:
        print(f"Filas con Filename en Excel: {summary.excel_rows}")
        print(f"Filename con match exacto: {summary.exact_matches}")
        print(f"Filename sin audio detectado: {summary.missing_in_audio_folder}")
        print(f"Audios sin fila en Excel: {summary.audio_without_excel_match}")
        print(f"Ratio de match exacto: {summary.match_ratio:.1%}")
        if summary.excel_rows and summary.match_ratio < 0.95:
            print("[WARN] Match exacto bajo. Revisa rutas relativas, subcarpetas, mayúsculas y slashes.")
    print()
    return summary


def normalize_text(text: str, transforms=wer_transforms) -> str:
    """Normalize text with a reusable jiwer transform pipeline."""
    return transforms(text or "")


def is_asian_language(lang: str) -> bool:
    """Return True for languages where space-tokenized WER is not meaningful."""
    if not isinstance(lang, str):
        return False
    return lang.strip().upper() in ASIAN_LANGUAGES


def calculate_wer(expected: str, transcript: str, language: Optional[str] = None) -> Optional[float]:
    """Calculate normalized WER, skipping languages without space-separated words."""
    if is_asian_language(language):
        return None

    expected_norm = normalize_text(expected, wer_transforms)
    transcript_norm = normalize_text(transcript, wer_transforms)
    if not expected_norm or not transcript_norm:
        return None

    return jiwer.wer(expected_norm, transcript_norm)


def calculate_cer(expected: str, transcript: str) -> Optional[float]:
    """Calculate minimally normalized CER."""
    expected_norm = normalize_text(expected, cer_transforms)
    transcript_norm = normalize_text(transcript, cer_transforms)
    if not expected_norm or not transcript_norm:
        return None

    return jiwer.cer(expected_norm, transcript_norm)


def classify_transcription_quality(wer_score: Optional[float], cer_score: Optional[float]) -> str:
    """Classify STT quality using centralized thresholds."""
    if wer_score is not None:
        if wer_score <= WER_EXCELLENT_THRESHOLD:
            return QUALITY_EXCELLENT
        if wer_score <= WER_ACCEPTABLE_THRESHOLD:
            return QUALITY_ACCEPTABLE
        return QUALITY_POOR

    if cer_score is not None:
        if cer_score <= CER_EXCELLENT_THRESHOLD:
            return QUALITY_EXCELLENT
        if cer_score <= CER_ACCEPTABLE_THRESHOLD:
            return QUALITY_ACCEPTABLE
        return QUALITY_POOR

    return QUALITY_UNEVALUATED


def format_score(score: Optional[float]) -> Optional[float]:
    """Return an Excel-friendly percentage value rounded to two decimals."""
    if score is None:
        return None
    return round(score * 100, 2)


def quality_fill(quality: str) -> PatternFill:
    if quality == QUALITY_EXCELLENT:
        return GREEN_FILL
    if quality == QUALITY_POOR:
        return RED_FILL
    return YELLOW_FILL


def transcribe_folder(
    transcriber: BaseTranscriber,
    folder_path: str,
    language: Optional[str] = "es",
) -> Dict[str, TranscriptionRecord]:
    files = get_audio_files(folder_path)
    if not files:
        print("[WARN] No se encontraron archivos de audio.")
        return {}

    transcripts: Dict[str, TranscriptionRecord] = {}

    # Deepgram: procesamiento concurrente para evitar espera secuencial archivo por archivo.
    if isinstance(transcriber, DeepgramTranscriber):
        max_workers = min(transcriber.max_workers, len(files))
        print(f"[INFO] Procesando {len(files)} audios con Deepgram en paralelo (workers={max_workers})")

        def _worker(audio_path: str) -> TranscriptionRecord:
            relative_path = os.path.relpath(audio_path, folder_path)
            text = transcriber.transcribe(audio_path, language=language)
            status = STATUS_OK if text else STATUS_EMPTY_TRANSCRIPT
            return TranscriptionRecord(audio_file=relative_path, transcript=text, status=status)

        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            future_map = {executor.submit(_worker, audio_path): audio_path for audio_path in files}
            done = 0
            total = len(files)
            for future in as_completed(future_map):
                audio_path = future_map[future]
                relative_path = os.path.relpath(audio_path, folder_path)
                done += 1
                try:
                    record = future.result()
                    transcripts[record.audio_file] = record
                    label = "OK" if record.status == STATUS_OK else "EMPTY"
                    print(f"[{done}/{total}] {label}: {record.audio_file}")
                except Exception as exc:
                    error_text = str(exc)
                    print(f"[{done}/{total}] [ERROR] Fallo {relative_path}: {error_text}")
                    transcripts[relative_path] = TranscriptionRecord(
                        audio_file=relative_path,
                        transcript="",
                        status=STATUS_FAILED,
                        error=error_text,
                    )

        return dict(sorted(transcripts.items()))

    # Whisper (u otros): secuencial para evitar problemas de memoria/contención de GPU.
    for idx, audio_path in enumerate(files, start=1):
        relative_path = os.path.relpath(audio_path, folder_path)
        print(f"[{idx}/{len(files)}] Transcribiendo: {relative_path}")
        try:
            text = transcriber.transcribe(audio_path, language=language)
            status = STATUS_OK if text else STATUS_EMPTY_TRANSCRIPT
            transcripts[relative_path] = TranscriptionRecord(
                audio_file=relative_path,
                transcript=text,
                status=status,
            )
        except Exception as exc:
            error_text = str(exc)
            print(f"[ERROR] Fallo {relative_path}: {error_text}")
            transcripts[relative_path] = TranscriptionRecord(
                audio_file=relative_path,
                transcript="",
                status=STATUS_FAILED,
                error=error_text,
            )

    return transcripts


def summarize_transcriptions(records: TranscriptMap) -> Counter:
    summary = Counter()
    for key, value in records.items():
        record = _record_from_value(key, value)
        summary[record.status] += 1
    print("\n=== Resumen de transcripción ===")
    print(f"Total audios detectados/procesados: {sum(summary.values())}")
    print(f"OK: {summary[STATUS_OK]}")
    print(f"Fallidos: {summary[STATUS_FAILED]}")
    print(f"Vacíos: {summary[STATUS_EMPTY_TRANSCRIPT]}")
    return summary

def export_transcriptions_to_excel(
    transcripts: TranscriptMap,
    output_path: str,
) -> List[TranscriptionOnlyResult]:
    # Si el usuario pasa una carpeta en lugar de archivo, agregamos un nombre por defecto
    if os.path.isdir(output_path):
        output_path = os.path.join(output_path, "resultado.xlsx")

    # Aseguramos que tenga la extension .xlsx
    if not output_path.lower().endswith(".xlsx"):
        output_path += ".xlsx"

    rows: List[TranscriptionOnlyResult] = []
    for audio_file, value in sorted(transcripts.items()):
        record = _record_from_value(audio_file, value)
        rows.append(
            TranscriptionOnlyResult(
                audio_file=audio_file,
                transcript=record.transcript,
                transcription_status=record.status,
                transcription_error=record.error,
            )
        )

    df = pd.DataFrame([
        {
            "audio_file": row.audio_file,
            "transcripcion": row.transcript,
            "transcription_status": row.transcription_status,
            "transcription_error": row.transcription_error,
        }
        for row in rows
    ])

    df.to_excel(output_path, index=False)
    return rows


def compare_with_excel(
    excel_path: str,
    transcripts: TranscriptMap,
    audio_column: str,
    expected_column: str,
    output_path: str,
    sheet_name: Optional[str] = None,
    language: Optional[str] = None,
) -> List[ComparisonResult]:
    # Si el usuario pasa una carpeta en lugar de archivo, agregamos un nombre por defecto
    if os.path.isdir(output_path):
        output_path = os.path.join(output_path, "resultado_comparacion.xlsx")

    # Aseguramos que tenga la extension .xlsx
    if not output_path.lower().endswith(".xlsx"):
        output_path += ".xlsx"

    # Leemos Excel de entrada
    df = pd.read_excel(excel_path, sheet_name=sheet_name)
    if isinstance(df, dict):
        if len(df) > 1:
            print("[WARN] Se detectaron multiples hojas, usando la primera.")
        df = list(df.values())[0]

    if audio_column not in df.columns or expected_column not in df.columns:
        raise ValueError(
            f"Columnas invalidas. Encontradas: {list(df.columns)} | "
            f"Esperadas: '{audio_column}' y '{expected_column}'"
        )

    normalized_records = {
        audio_file: _record_from_value(audio_file, value)
        for audio_file, value in transcripts.items()
    }
    matched_audio: set[str] = set()

    results: List[ComparisonResult] = []
    transcript_list: List[str] = []
    wer_list: List[Optional[float]] = []
    cer_list: List[Optional[float]] = []
    quality_list: List[str] = []
    status_list: List[str] = []
    error_list: List[str] = []

    for _, row in df.iterrows():
        audio_name = str(row[audio_column]).strip()
        expected_text = str(row[expected_column]) if pd.notna(row[expected_column]) else ""

        record = normalized_records.get(audio_name)
        if record is None:
            record = TranscriptionRecord(
                audio_file=audio_name,
                transcript="",
                status=STATUS_MISSING_AUDIO,
                error="No se encontró un audio/transcripción con Filename exacto.",
            )
        else:
            matched_audio.add(audio_name)

        transcript = record.transcript
        wer_score = calculate_wer(expected_text, transcript, language=language)
        cer_score = calculate_cer(expected_text, transcript)
        quality = classify_transcription_quality(wer_score, cer_score)

        results.append(
            ComparisonResult(
                audio_file=audio_name,
                transcript=transcript,
                expected_text=expected_text,
                wer=wer_score,
                cer=cer_score,
                quality=quality,
                transcription_status=record.status,
                transcription_error=record.error,
            )
        )
        transcript_list.append(transcript)
        wer_list.append(format_score(wer_score))
        cer_list.append(format_score(cer_score))
        quality_list.append(quality)
        status_list.append(record.status)
        error_list.append(record.error)

    df["transcripcion"] = transcript_list
    df["wer_pct"] = wer_list
    df["cer_pct"] = cer_list
    df["quality"] = quality_list
    df["transcription_status"] = status_list
    df["transcription_error"] = error_list
    df.to_excel(output_path, index=False)

    # Pintar en colores usando openpyxl y agregar hoja con audios que no matchearon Excel.
    wb = load_workbook(output_path)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    wer_col_idx = list(df.columns).index("wer_pct") + 1
    cer_col_idx = list(df.columns).index("cer_pct") + 1
    quality_col_idx = list(df.columns).index("quality") + 1
    status_col_idx = list(df.columns).index("transcription_status") + 1

    for row_idx, quality in enumerate(quality_list, start=2):
        fill = quality_fill(quality)

        for col_idx in (wer_col_idx, cer_col_idx, quality_col_idx, status_col_idx):
            ws.cell(row=row_idx, column=col_idx).fill = fill

    unmatched = [
        record
        for audio_file, record in sorted(normalized_records.items())
        if audio_file not in matched_audio
    ]
    if unmatched:
        if "unmatched_audio" in wb.sheetnames:
            del wb["unmatched_audio"]
        unmatched_ws = wb.create_sheet("unmatched_audio")
        headers = ["audio_file", "transcripcion", "transcription_status", "transcription_error"]
        unmatched_ws.append(headers)
        for record in unmatched:
            unmatched_ws.append([
                record.audio_file,
                record.transcript,
                STATUS_NO_EXCEL_MATCH,
                record.error or "Audio detectado/transcrito, pero no existe Filename exacto en Excel.",
            ])

    wb.save(output_path)

    compare_summary = Counter(status_list)
    print("\n=== Resumen Compare ===")
    print(f"Total filas Excel: {len(results)}")
    print(f"OK: {compare_summary[STATUS_OK]}")
    print(f"Fallidos: {compare_summary[STATUS_FAILED]}")
    print(f"Vacíos: {compare_summary[STATUS_EMPTY_TRANSCRIPT]}")
    print(f"Audios faltantes/no detectados: {compare_summary[STATUS_MISSING_AUDIO]}")
    print(f"Audios detectados sin match en Excel: {len(unmatched)}")

    return results

def export_qa_reaper_project(
    audio_folder: str,
    results: List[ComparisonResult],
    output_path: str,
) -> str:
    """Export comparison results as a REAPER QA project grouped by quality."""
    rpp_path = Path(output_path)
    if rpp_path.suffix.lower() != ".rpp":
        rpp_path = rpp_path.with_suffix(".rpp")

    generate_qa_project(
        QAProjectConfig(
            source_root=Path(audio_folder),
            output_file=rpp_path,
        ),
        results,
    )
    return str(rpp_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Transcribe audios y opcionalmente compara contra un Excel de guion."
    )
    parser.add_argument(
        "--mode",
        choices=["compare", "transcribe-only"],
        default="compare",
        help="compare: compara contra Excel. transcribe-only: solo transcribe y exporta a nuevo Excel.",
    )
    parser.add_argument("--audio-folder", required=False, help="Carpeta con audios")
    parser.add_argument("--excel", required=False, help="Ruta del Excel de entrada (solo modo compare)")
    parser.add_argument(
        "--audio-column",
        default="Filename",
        help="Nombre de columna con nombre/ruta relativa del audio",
    )
    parser.add_argument(
        "--expected-column",
        default="Script",
        help="Nombre de columna con texto esperado",
    )
    parser.add_argument("--sheet", default=None, help="Nombre de hoja en Excel (opcional)")
    parser.add_argument(
        "--output",
        default="resultado_comparacion.xlsx",
        help="Ruta del Excel de salida",
    )
    # Busca esta sección en parse_args()
    parser.add_argument(
        "--engine",
        choices=["whisper", "deepgram"],
        default="deepgram",
        help="Motor de transcripción: deepgram o whisper",
    )
    parser.add_argument(
        "--model-size",
        default="medium",
        help="Modelo Whisper (tiny/base/small/medium/large/large-v3)",
    )
    parser.add_argument(
        "--deepgram-model",
        default="nova-3",
        help="Modelo Deepgram (ej. nova-3)",
    )
    parser.add_argument(
        "--language",
        default="es",
        help="Idioma forzado (ej. es, en) o vacio para autodetectar",
    )
    parser.add_argument(
        "--glossary",
        default=None,
        help="Excel de glosario para Deepgram (columnas: term, boost, variants, enabled, notes).",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Solo valida audios/Excel y no llama a ningún motor STT.",
    )
    parser.add_argument(
        "--log-file",
        default=None,
        help="Ruta opcional para guardar el log. Por defecto crea asrec_run_YYYY-MM-DD_HHMM.log junto al output.",
    )
    parser.add_argument(
        "--deepgram-retries",
        type=int,
        default=int(os.getenv("DEEPGRAM_MAX_RETRIES", str(DEFAULT_DEEPGRAM_RETRIES))),
        help="Cantidad máxima de intentos por audio para errores transitorios de Deepgram.",
    )
    parser.add_argument(
        "--deepgram-retry-backoff",
        type=float,
        default=float(os.getenv("DEEPGRAM_RETRY_BACKOFF_SECONDS", str(DEFAULT_DEEPGRAM_RETRY_BACKOFF_SECONDS))),
        help="Espera base en segundos para backoff exponencial de Deepgram.",
    )

    return parser.parse_args()


def ask_input(prompt: str, default: Optional[str] = None) -> str:
    if default is not None:
        value = input(f"{prompt} [{default}]: ").strip()
        if value == "":
            value = default
    else:
        value = input(f"{prompt}: ").strip()
    return value.strip().strip('"').strip("'")


def main() -> None:
    args = parse_args()

    # Detectar doble click / sin argumentos
    interactive_mode = len(sys.argv) == 1

    if interactive_mode:
        print("\n=== MODO INTERACTIVO ===\n")

        args.mode = ask_input(
            "Modo (1=compare, 2=transcribe-only)",
            "1"
        )

        args.mode = "compare" if args.mode == "1" else "transcribe-only"

        args.audio_folder = ask_input("Ruta de carpeta de audios")

        if args.mode == "compare":
            args.excel = ask_input("Ruta del Excel")

            args.audio_column = ask_input(
                "Columna de audio",
                args.audio_column
            )

            args.expected_column = ask_input(
                "Columna de texto esperado",
                args.expected_column
            )

            args.sheet = ask_input(
                "Nombre de hoja (Enter = primera)",
                ""
            ) or None

            args.output = ask_input(
                "Ruta de salida",
                "resultado_comparacion.xlsx"
            )

        else:  # transcribe-only
            args.output = ask_input(
                "Ruta de salida",
                "solo_transcripciones.xlsx"
            )

        args.engine = ask_input(
            "Motor (whisper/deepgram)",
            args.engine
        ).lower()

        if args.engine == "whisper":
            args.model_size = ask_input(
                "Modelo Whisper",
                args.model_size
            )
        else:
            args.deepgram_model = ask_input(
                "Modelo Deepgram",
                args.deepgram_model
            )
            args.glossary = ask_input(
                "Ruta del glosario Deepgram (Enter para omitir)",
                ""
            ) or None

        args.language = ask_input(
            "Idioma (es/en/... o vacio auto)",
            args.language
        )

    # Validacion minima (por si viene por CLI)
    if not args.audio_folder:
        args.audio_folder = ask_input("Ruta de carpeta de audios")

    if args.mode == "compare" and not args.excel:
        args.excel = ask_input("Ruta del Excel")

    if not args.output:
        args.output = ask_input("Ruta de salida", "resultado.xlsx")

    with run_logging(args.log_file, args.output) as log_path:
        print("\n[INFO] Configuracion final:")
        print(f"Modo: {args.mode}")
        print(f"Audios: {args.audio_folder}")
        print(f"Output: {args.output}")
        print(f"Dry run: {args.dry_run}")
        print()

        validate_inputs(
            audio_folder=args.audio_folder,
            excel_path=args.excel if args.mode == "compare" else None,
            audio_column=args.audio_column,
            sheet_name=args.sheet,
        )

        if args.dry_run:
            print("[INFO] Dry run completado: no se llamó a ningún motor STT.")
            print(f"Log: {log_path}")
            if interactive_mode:
                input("\nPresiona Enter para cerrar...")
            return

        # Ejecución
        language = args.language if args.language else None

        if args.engine == "deepgram":
            deepgram_api_key = os.getenv("DEEPGRAM_API_KEY", "").strip()
            deepgram_workers = int(os.getenv("DEEPGRAM_MAX_WORKERS", "4"))
            transcriber = DeepgramTranscriber(
                api_key=deepgram_api_key,
                model=args.deepgram_model,
                max_workers=deepgram_workers,
                glossary_path=args.glossary,
                max_retries=args.deepgram_retries,
                retry_backoff_seconds=args.deepgram_retry_backoff,
            )
        else:
            transcriber = WhisperTranscriber(model_size=args.model_size)

        transcripts = transcribe_folder(
            transcriber=transcriber,
            folder_path=args.audio_folder,
            language=language,
        )
        summarize_transcriptions(transcripts)

        if args.mode == "transcribe-only":
            rows = export_transcriptions_to_excel(
                transcripts=transcripts,
                output_path=args.output
            )

            print("\n=== Resumen (solo transcripcion) ===")
            print(f"Audios procesados: {len(rows)}")
            print(f"Salida: {args.output}")
            print(f"Log: {log_path}")

            if interactive_mode:
                input("\nPresiona Enter para cerrar...")

            return

        # Modo compare
        results = compare_with_excel(
            excel_path=args.excel,
            transcripts=transcripts,
            audio_column=args.audio_column,
            expected_column=args.expected_column,
            output_path=args.output,
            sheet_name=args.sheet,
            language=language,
        )

        total = len(results)
        excellent = sum(1 for x in results if x.quality == QUALITY_EXCELLENT)
        acceptable = sum(1 for x in results if x.quality == QUALITY_ACCEPTABLE)
        poor = sum(1 for x in results if x.quality == QUALITY_POOR)
        unevaluated = sum(1 for x in results if x.quality == QUALITY_UNEVALUATED)

        print("\n=== Resumen de calidad ===")
        print(f"Total filas: {total}")
        print(f"Excelente (verde): {excellent}")
        print(f"Aceptable (amarillo): {acceptable}")
        print(f"Pobre (rojo): {poor}")
        print(f"Sin evaluar (amarillo): {unevaluated}")
        print(f"Salida: {args.output}")
        print(f"Log: {log_path}")

        if interactive_mode:
            input("\nPresiona Enter para cerrar...")


if __name__ == "__main__":
    main()
