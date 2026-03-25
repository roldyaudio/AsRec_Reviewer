import argparse
import os
import sys
import re
import unicodedata
from dataclasses import dataclass
from typing import Dict, Iterable, List, Optional

import pandas as pd
import torch
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import difflib


# Colores para Excel
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")


class BaseTranscriber:
    """Interfaz base para cambiar fácilmente el backend de transcripción."""

    def transcribe(self, audio_path: str, language: Optional[str] = None) -> str:
        raise NotImplementedError


class WhisperTranscriber(BaseTranscriber):
    """Implementación con OpenAI Whisper local."""

    def __init__(self, model_size: str = "medium"):
        import whisper

        self.device = "cuda" if torch.cuda.is_available() else "cpu"
        self.model_size = model_size
        print(f"[INFO] Usando dispositivo: {self.device}")
        print(f"[INFO] Cargando modelo Whisper: {model_size}")
        self.model = whisper.load_model(model_size, device=self.device)

    def transcribe(self, audio_path: str, language: Optional[str] = None) -> str:
        kwargs = {
            "fp16": self.device == "cuda",
        }
        if language:
            kwargs["language"] = language

        result = self.model.transcribe(audio_path, **kwargs)
        return result.get("text", "").strip()


@dataclass
class ComparisonResult:
    audio_file: str
    transcript: str
    expected_text: str
    match: Optional[bool]  # True/False/None(no evaluable)


@dataclass
class TranscriptionOnlyResult:
    audio_file: str
    transcript: str


def get_audio_files(folder_path: str, extensions: Iterable[str] = (".wav", ".mp3", ".m4a")) -> List[str]:
    audio_files: List[str] = []
    allowed = tuple(ext.lower() for ext in extensions)

    for root, _, files in os.walk(folder_path):
        for file in files:
            if file.lower().endswith(allowed):
                audio_files.append(os.path.join(root, file))

    audio_files.sort()
    return audio_files


def normalize_text(text: str) -> str:
    """
    Normaliza para comparación robusta:
    - minúsculas
    - elimina tildes
    - quita puntuación
    - compacta espacios
    """
    text = text or ""
    text = text.lower().strip()
    text = "".join(
        c for c in unicodedata.normalize("NFD", text) if unicodedata.category(c) != "Mn"
    )
    text = re.sub(r"[^\w\s]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def exact_or_contains_match(expected: str, transcript: str) -> Optional[bool]:
    """
    Regla simple:
    - None si falta texto esperado o transcrito
    - True si coinciden normalizados o uno contiene al otro
    - False en cualquier otro caso
    """
    exp_norm = normalize_text(expected)
    tr_norm = normalize_text(transcript)

    if not exp_norm or not tr_norm:
        return None

    if exp_norm == tr_norm:
        return True

    if exp_norm in tr_norm or tr_norm in exp_norm:
        return True

    return False


def transcribe_folder(
    transcriber: BaseTranscriber,
    folder_path: str,
    language: Optional[str] = "es",
) -> Dict[str, str]:
    files = get_audio_files(folder_path)
    if not files:
        print("[WARN] No se encontraron archivos de audio.")
        return {}

    transcripts: Dict[str, str] = {}
    for idx, audio_path in enumerate(files, start=1):
        relative_path = os.path.relpath(audio_path, folder_path)
        print(f"[{idx}/{len(files)}] Transcribiendo: {relative_path}")
        try:
            transcripts[relative_path] = transcriber.transcribe(audio_path, language=language)
        except Exception as exc:
            print(f"[ERROR] Falló {relative_path}: {exc}")
            transcripts[relative_path] = ""

    return transcripts


def highlight_differences(expected: str, transcript: str) -> str:
    """
    Genera una versión del texto donde las diferencias se marcan con *palabra*.
    """
    exp_words = expected.split()
    tr_words = transcript.split()
    s = difflib.SequenceMatcher(None, exp_words, tr_words)

    highlighted = []
    for opcode, _i1, _i2, j1, j2 in s.get_opcodes():
        if opcode == "equal":
            highlighted.extend(tr_words[j1:j2])
        elif opcode in ("replace", "delete", "insert"):
            highlighted.extend([f"*{w}*" for w in tr_words[j1:j2]])
    return " ".join(highlighted)


def export_transcriptions_to_excel(
    transcripts: Dict[str, str],
    output_path: str,
) -> List[TranscriptionOnlyResult]:
    if os.path.isdir(output_path):
        output_path = os.path.join(output_path, "resultado.xlsx")

    if not output_path.lower().endswith(".xlsx"):
        output_path += ".xlsx"

    rows: List[TranscriptionOnlyResult] = [
        TranscriptionOnlyResult(audio_file=audio_file, transcript=transcript)
        for audio_file, transcript in sorted(transcripts.items())
    ]

    df = pd.DataFrame([
        {"audio_file": row.audio_file, "transcripcion": row.transcript}
        for row in rows
    ])

    df.to_excel(output_path, index=False)

    return rows


def compare_with_excel(
    excel_path: str,
    transcripts: Dict[str, str],
    audio_column: str,
    expected_column: str,
    output_path: str,
    sheet_name: Optional[str] = None,
) -> List[ComparisonResult]:
    if os.path.isdir(output_path):
        output_path = os.path.join(output_path, "resultado_comparacion.xlsx")

    if not output_path.lower().endswith(".xlsx"):
        output_path += ".xlsx"

    df = pd.read_excel(excel_path, sheet_name=sheet_name)
    if isinstance(df, dict):
        print("[WARN] Se detectaron múltiples hojas, usando la primera.")
        df = list(df.values())[0]

    if audio_column not in df.columns or expected_column not in df.columns:
        raise ValueError(
            f"Columnas inválidas. Encontradas: {list(df.columns)} | "
            f"Esperadas: '{audio_column}' y '{expected_column}'"
        )

    results: List[ComparisonResult] = []
    transcript_list: List[str] = []
    match_list: List[Optional[bool]] = []

    for _, row in df.iterrows():
        audio_name = str(row[audio_column]).strip()
        expected_text = str(row[expected_column]) if pd.notna(row[expected_column]) else ""

        transcript = transcripts.get(audio_name, "")
        match = exact_or_contains_match(expected_text, transcript)

        results.append(
            ComparisonResult(
                audio_file=audio_name,
                transcript=transcript,
                expected_text=expected_text,
                match=match,
            )
        )
        transcript_list.append(transcript)
        match_list.append(match)

    df["transcripcion"] = transcript_list
    df["coincide"] = match_list
    df.to_excel(output_path, index=False)

    wb = load_workbook(output_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    expected_col_idx = list(df.columns).index(expected_column) + 1
    transcript_col_idx = list(df.columns).index("transcripcion") + 1
    match_col_idx = list(df.columns).index("coincide") + 1

    for row_idx, value in enumerate(match_list, start=2):
        expected_cell = ws.cell(row=row_idx, column=expected_col_idx)
        transcript_cell = ws.cell(row=row_idx, column=transcript_col_idx)
        match_cell = ws.cell(row=row_idx, column=match_col_idx)

        if value is True:
            expected_cell.fill = GREEN_FILL
            transcript_cell.fill = GREEN_FILL
            match_cell.fill = GREEN_FILL
        elif value is False:
            expected_cell.fill = RED_FILL
            transcript_cell.value = highlight_differences(str(expected_cell.value or ""), str(transcript_cell.value or ""))
            transcript_cell.fill = RED_FILL
            match_cell.fill = RED_FILL
        else:
            expected_cell.fill = YELLOW_FILL
            transcript_cell.fill = YELLOW_FILL
            match_cell.fill = YELLOW_FILL

    wb.save(output_path)
    return results


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Transcribe audios y opcionalmente compara contra un Excel de guion."
    )
    parser.add_argument(
        "--mode",
        choices=["compare", "transcribe-only"],
        default="compare",
        help="compare: compara contra Excel. transcribe-only: solo transcribe y exporta a nuevo Excel.",
    )
    parser.add_argument("--audio-folder", required=False, help="Carpeta con audios")
    parser.add_argument("--excel", required=False, help="Ruta del Excel de entrada (solo modo compare)")
    parser.add_argument(
        "--audio-column",
        default="audio_file",
        help="Nombre de columna con nombre/ruta relativa del audio",
    )
    parser.add_argument(
        "--expected-column",
        default="guion",
        help="Nombre de columna con texto esperado",
    )
    parser.add_argument("--sheet", default=None, help="Nombre de hoja en Excel (opcional)")
    parser.add_argument(
        "--output",
        default="resultado_comparacion.xlsx",
        help="Ruta del Excel de salida",
    )
    parser.add_argument(
        "--model-size",
        default="medium",
        help="Modelo Whisper (tiny/base/small/medium/large)",
    )
    parser.add_argument(
        "--language",
        default="es",
        help="Idioma forzado (ej. es, en) o vacío para autodetectar",
    )

    return parser.parse_args()


def ask_input(prompt: str, default: Optional[str] = None) -> str:
    if default is not None:
        value = input(f"{prompt} [{default}]: ").strip()
        if value == "":
            value = default
    else:
        value = input(f"{prompt}: ").strip()
    return value.strip().strip('"').strip("'")


def main() -> None:
    args = parse_args()

    interactive_mode = len(sys.argv) == 1

    if interactive_mode:
        print("\n=== MODO INTERACTIVO ===\n")

        args.mode = ask_input(
            "Modo (1=compare, 2=transcribe-only)",
            "1"
        )

        args.mode = "compare" if args.mode == "1" else "transcribe-only"

        args.audio_folder = ask_input("Ruta de carpeta de audios")

        if args.mode == "compare":
            args.excel = ask_input("Ruta del Excel")

            args.audio_column = ask_input(
                "Columna de audio",
                args.audio_column
            )

            args.expected_column = ask_input(
                "Columna de texto esperado",
                args.expected_column
            )

            args.sheet = ask_input(
                "Nombre de hoja (Enter = primera)",
                ""
            ) or None

            args.output = ask_input(
                "Ruta de salida",
                "resultado_comparacion.xlsx"
            )

        else:
            args.output = ask_input(
                "Ruta de salida",
                "solo_transcripciones.xlsx"
            )

        args.model_size = ask_input(
            "Modelo Whisper",
            args.model_size
        )

        args.language = ask_input(
            "Idioma (es/en/... o vacío auto)",
            args.language
        )

    if not args.audio_folder:
        args.audio_folder = ask_input("Ruta de carpeta de audios")

    if args.mode == "compare" and not args.excel:
        args.excel = ask_input("Ruta del Excel")

    if not args.output:
        args.output = ask_input("Ruta de salida", "resultado.xlsx")

    print("\n[INFO] Configuración final:")
    print(f"Modo: {args.mode}")
    print(f"Audios: {args.audio_folder}")
    print(f"Output: {args.output}")
    print()

    transcriber = WhisperTranscriber(model_size=args.model_size)
    language = args.language if args.language else None

    transcripts = transcribe_folder(
        transcriber=transcriber,
        folder_path=args.audio_folder,
        language=language,
    )

    if args.mode == "transcribe-only":
        rows = export_transcriptions_to_excel(
            transcripts=transcripts,
            output_path=args.output
        )

        print("\n=== Resumen (solo transcripción) ===")
        print(f"Audios procesados: {len(rows)}")
        print(f"Salida: {args.output}")

        if interactive_mode:
            input("\nPresiona Enter para cerrar...")

        return

    results = compare_with_excel(
        excel_path=args.excel,
        transcripts=transcripts,
        audio_column=args.audio_column,
        expected_column=args.expected_column,
        output_path=args.output,
        sheet_name=args.sheet,
    )

    total = len(results)
    ok = sum(1 for x in results if x.match is True)
    bad = sum(1 for x in results if x.match is False)
    unk = sum(1 for x in results if x.match is None)

    print("\n=== Resumen ===")
    print(f"Total filas: {total}")
    print(f"Coinciden (verde): {ok}")
    print(f"No coinciden (rojo): {bad}")
    print(f"Sin evaluar (amarillo): {unk}")
    print(f"Salida: {args.output}")

    if interactive_mode:
        input("\nPresiona Enter para cerrar...")


if __name__ == "__main__":
    main()
